import csv
import os.path
import struct
import subprocess
import sys

import pandas as pd
from PyQt5.QtCore import Qt, QAbstractTableModel, QTime
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableView, QPushButton, QHBoxLayout, QVBoxLayout, QWidget, \
    QLabel, QHeaderView, QStyledItemDelegate, QSpinBox, QAction, QTimeEdit, QComboBox, QStackedWidget, QStyle
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from settings import *


class SpinBoxDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QSpinBox(parent)
        editor.setMinimum(0)
        editor.setMaximum(100)
        editor.setAlignment(Qt.AlignCenter)
        return editor

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.EditRole)
        editor.setValue(int(value))

    def setModelData(self, editor, model, index):
        value = editor.value()
        model.setData(index, value, Qt.EditRole)


class PandasModel(QAbstractTableModel):
    def __init__(self, data: pd.DataFrame):
        super().__init__()
        self._data = data

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        col = index.column()
        row = index.row()

        if col == 0:
            if role == Qt.CheckStateRole:
                return Qt.Checked if self._data.iat[row, col] else Qt.Unchecked
            return None

        if role in (Qt.DisplayRole, Qt.EditRole):
            return str(self._data.iat[row, col])
        return None

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                colName = str(self._data.columns[section])
                if colName == 'required':
                    return ''
                return colName
        return None

    def flags(self, index):
        flags = super().flags(index)
        if index.column() == 0:
            flags |= Qt.ItemIsUserCheckable
        elif index.column() == 1:
            flags |= Qt.ItemIsEditable
        return flags

    def setData(self, index, value, role):
        col = index.column()
        row = index.row()

        if col == 0 and role == Qt.CheckStateRole:
            self._data.iat[row, col] = (value == Qt.Checked)
            self.dataChanged.emit(index, index)
            return True
        elif col == 1 and role == Qt.EditRole:
            try:
                int_value = int(value)
                if 0 <= int_value <= 100:
                    self._data.iat[row, col] = int_value
                    self.dataChanged.emit(index, index)
                    return True
            except ValueError:
                pass

        return False


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Starter')
        self.setWindowIcon(QIcon('icon/plane.png'))
        self.level1Sheets = None
        self.level2Sheets = None
        self.level3Sheets = None
        self.level4Sheets = None
        self.timePicker = None
        self.weatherPicker = None
        self.loadData()
        self.initUi()

    def loadData(self):
        if os.path.exists(LEVEL_1_EXCEL):
            self.level1Sheets = pd.read_excel(LEVEL_1_EXCEL, sheet_name=None, dtype={'id': str})
        if os.path.exists(LEVEL_2_EXCEL):
            self.level2Sheets = pd.read_excel(LEVEL_2_EXCEL, sheet_name=None, dtype={'id': str})
        if os.path.exists(LEVEL_3_EXCEL):
            self.level3Sheets = pd.read_excel(LEVEL_3_EXCEL, sheet_name=None, dtype={'id': str})
        if os.path.exists(LEVEL_4_EXCEL):
            self.level4Sheets = pd.read_excel(LEVEL_4_EXCEL, sheet_name=None, dtype={'id': str})

        for level in [self.level1Sheets, self.level2Sheets,
                      self.level3Sheets, self.level4Sheets]:
            if level is None:
                continue
            for name, df in level.items():
                df['id'] = df['id'].ffill()
                if 'Question' in df.columns:
                    df = (
                        df.groupby('id', as_index=False)
                        .agg({
                            'Question': ' '.join
                        })
                    )

                df.insert(0, 'required', True)
                df.insert(1, 'score', 10)
                level[name] = df

    def createAction(self, name, level):
        if level == 0:
            action = QAction(QIcon('icon/cloud.png'), 'Environment', self)
            action.triggered.connect(lambda _, i=level: self.stacked.setCurrentIndex(i))
            self.toolbar.addAction(action)
            return
        elif level == 1:
            action = QAction(QIcon('icon/one.png'), 'Level1', self)
        elif level == 2:
            action = QAction(QIcon('icon/two.png'), 'Level2', self)
        elif level == 3:
            action = QAction(QIcon('icon/three.png'), 'Level3', self)
        elif level == 4:
            action = QAction(QIcon('icon/four.png'), 'Level4', self)
        action.triggered.connect(lambda _, i=level: self.stacked.setCurrentIndex(i))
        self.toolbar.addAction(action)

    def createEnvPage(self):
        widget = QWidget()
        envLabel = QLabel(LEVEL_0_NAME)
        envLabel.setStyleSheet("font-weight:bold; margin-left:0px; margin-top:10px; margin-bottom:10px")

        timehBox = QHBoxLayout()
        timeLabel = QLabel('时间：')
        self.timePicker = QTimeEdit()
        self.timePicker.setTime(QTime(12, 0))
        timehBox.addWidget(timeLabel)
        timehBox.addWidget(self.timePicker)
        timehBox.addStretch(2)

        weatherhBox = QHBoxLayout()
        weatherLabel = QLabel('天气：')
        self.weatherPicker = QComboBox()
        self.weatherPicker.addItems(['晴', '多云', '雾', '打雷', '雷暴', '雪', '雨'])
        weatherhBox.addWidget(weatherLabel)
        weatherhBox.addWidget(self.weatherPicker)
        weatherhBox.addStretch(2)

        vbox = QVBoxLayout(widget)
        vbox.addWidget(envLabel)
        vbox.addLayout(timehBox)
        vbox.addLayout(weatherhBox)
        vbox.addStretch(3)

        return widget

    def createLevelPage(self, level):
        widget = QWidget()

        if level == 1:
            sheets = self.level1Sheets
        elif level == 2:
            sheets = self.level2Sheets
        elif level == 3:
            sheets = self.level3Sheets
        elif level == 4:
            sheets = self.level4Sheets

        vbox = QVBoxLayout(widget)

        if level == 1:
            levelLabel = QLabel(LEVEL_1_NAME)
        elif level == 2:
            levelLabel = QLabel(LEVEL_2_NAME)
        elif level == 3:
            levelLabel = QLabel(LEVEL_3_NAME)
        elif level == 4:
            levelLabel = QLabel(LEVEL_4_NAME)
        levelLabel.setStyleSheet("font-weight:bold; margin-left:0px; margin-top:10px; margin-bottom:10px")
        vbox.addWidget(levelLabel)

        if sheets is None:
            errorLabel = QLabel('本关卡题目配置读取失败，请确认路径是否正确！')
            vbox.addWidget(errorLabel)
            vbox.addStretch(2)
            return widget

        for sheet_name, df in sheets.items():
            if not df.empty:
                # 标题
                label = QLabel(f"{sheet_name}")
                vbox.addWidget(label)

                # 表格
                model = PandasModel(df)
                table = QTableView()
                table.setModel(model)
                table.resizeColumnsToContents()

                # 设置委托
                table.setItemDelegateForColumn(1, SpinBoxDelegate())

                for col in range(model.columnCount()):
                    header = model.headerData(col, Qt.Horizontal)
                    if header not in ('required', 'score', 'id', 'Question', 'Content'):
                        table.hideColumn(col)

                # 调整列宽
                table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
                table.showColumn(0)
                table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
                table.setColumnWidth(1, 80)
                table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
                table.setColumnWidth(2, 80)
                table.horizontalHeader().setStretchLastSection(True)

                vbox.addWidget(table)

        return widget

    def initUi(self):
        self.toolbar = self.addToolBar('')
        self.toolbar.setMovable(False)
        self.toolbar.setContextMenuPolicy(Qt.PreventContextMenu)
        self.createAction('Environment', 0)
        self.createAction('Level1', 1)
        self.createAction('Level2', 2)
        self.createAction('Level3', 3)
        self.createAction('Level4', 4)

        self.stacked = QStackedWidget()
        self.stacked.addWidget(self.createEnvPage())
        for i in (1, 2, 3, 4):
            self.stacked.addWidget(self.createLevelPage(i))

        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.okButtonClicked)

        style = QApplication.style()
        defaultLeft = style.pixelMetric(QStyle.PM_LayoutLeftMargin)
        defaultTop = style.pixelMetric(QStyle.PM_LayoutTopMargin)
        defaultRight = style.pixelMetric(QStyle.PM_LayoutRightMargin)
        defaultBottom = style.pixelMetric(QStyle.PM_LayoutBottomMargin)

        central = QWidget()
        layout = QVBoxLayout(central)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.stacked)
        bottom = QHBoxLayout()
        bottom.setContentsMargins(defaultLeft, defaultTop, defaultRight, defaultBottom)
        bottom.addStretch(1)
        bottom.addWidget(ok_btn)
        layout.addLayout(bottom)
        self.setCentralWidget(central)

    def okButtonClicked(self):
        time = int(str(self.timePicker.time().hour()).zfill(2) + str(self.timePicker.time().minute()).zfill(2))
        weather = self.weatherPicker.currentIndex()
        with open(BINARY_OUTPUT, 'wb') as file:
            packed = struct.pack('<HH', time, weather)
            file.write(packed)
        file.close()

        keepIds = []
        with open(QUESTION_OUTPUT, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for level in (self.level1Sheets, self.level2Sheets, self.level3Sheets, self.level4Sheets):
                if level is None:
                    continue
                for df in level.values():
                    for index, row in df.iterrows():
                        if row['required']:
                            writer.writerow([str(row['id'])] + [str(row['score'])])
                            keepIds.append(str(row['id']))
        csvfile.close()

        levelIndex = 0
        for path in (LEVEL_1_EXCEL, LEVEL_2_EXCEL, LEVEL_3_EXCEL, LEVEL_4_EXCEL):
            sheets = load_workbook(path)
            for ws in sheets.worksheets:
                ws = self.splitMergedCells(ws)
                rowsToCheck = list(ws.iter_rows(min_row=2, values_only=False))
                for cellTuple in reversed(rowsToCheck):
                    rowIdx = cellTuple[0].row
                    if cellTuple[0].value not in keepIds:
                        self.deleteRow(ws, rowIdx)
                ws = self.mergeIndeticalCells(ws)
                sheets.save(OUTPUT_EXCEL[levelIndex])
            levelIndex += 1

        if os.path.exists(TRACKER_APPLICATION):
            subprocess.Popen(TRACKER_APPLICATION)
        if os.path.exists(UNREAL_APPLICATION):
            subprocess.Popen(UNREAL_APPLICATION)

    def splitMergedCells(self, sheet: Worksheet):
        merged_info = []
        for merged_range in list(sheet.merged_cells.ranges):
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            max_row = merged_range.max_row
            max_col = merged_range.max_col

            top_left_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(start_row=min_row, start_column=min_col,
                                end_row=max_row, end_column=max_col)
            merged_info.append((min_row, min_col, max_row, max_col, top_left_value))

        for info in merged_info:
            min_row, min_col, max_row, max_col, value = info
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col, value=value)

        return sheet

    def mergeIndeticalCells(self, sheet: Worksheet):
        id_dict = {}
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1), start=1):
            cell_id = row[0].value
            if cell_id not in id_dict:
                id_dict[cell_id] = []
            id_dict[cell_id].append(row_idx)

        for group_id, row_indices in id_dict.items():
            if len(row_indices) < 2:
                continue
            for col_idx in range(1, sheet.max_column + 1):
                values = [
                    sheet.cell(row=row, column=col_idx).value
                    for row in row_indices
                ]

                if all(v == values[0] for v in values):
                    min_row = min(row_indices)
                    max_row = max(row_indices)

                    sheet.merge_cells(
                        start_row=min_row, start_column=col_idx,
                        end_row=max_row, end_column=col_idx
                    )
        return sheet

    def deleteRow(self, sheet: Worksheet, idx: int):
        to_remove = []
        for mcr in list(sheet.merged_cells.ranges):
            if idx == mcr.min_row:
                to_remove.append(mcr)
        sheet.delete_rows(idx)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.resize(1200, 1200)
    window.show()
    sys.exit(app.exec_())
